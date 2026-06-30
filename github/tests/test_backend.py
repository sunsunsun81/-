from __future__ import annotations

import io
import re
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfWriter
import zxingcpp

from server.app import create_app
from server.data_store import DEFAULT_ADMIN, DEFAULT_COMPANY_ID, DEFAULT_COMPANY_NAME, DEFAULT_PASSWORD, JsonStore
from server.env_check import check_runtime_environment, format_environment_report
from server.excel_service import create_template, read_invoice_rows
from server.invoice_parser import parse_invoice_file, parse_invoice_text


class BackendTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.store = JsonStore(Path(self.tmp.name))
        self.app = create_app(self.store)
        self.client = self.app.test_client()

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def login(self, username: str = DEFAULT_ADMIN, password: str = DEFAULT_PASSWORD) -> None:
        response = self.client.post("/api/auth/login", json={"username": username, "password": password})
        self.assertEqual(response.status_code, 200)

    def sample_invoice(self, code: str = "2026-06-24-001", number: str = "12345678") -> dict:
        return {
            "invoice_code": code,
            "original_invoice_code": "011002300111",
            "invoice_number": number,
            "issue_date": "2024-05-21",
            "buyer_name": "购买方",
            "seller_name": "销售方",
            "amount": "100.00",
            "tax_amount": "13.00",
            "total_amount": "113.00",
            "remark": "",
        }

    def make_pdf_bytes(self) -> bytes:
        writer = PdfWriter()
        writer.add_blank_page(width=240, height=160)
        buffer = io.BytesIO()
        writer.write(buffer)
        return buffer.getvalue()

    def make_jpg_bytes(self) -> bytes:
        image = Image.new("RGB", (360, 180), "white")
        buffer = io.BytesIO()
        image.save(buffer, format="JPEG")
        return buffer.getvalue()

    def make_qr_png_bytes(self, text: str) -> bytes:
        bitmap = zxingcpp.write_barcode(zxingcpp.BarcodeFormat.QRCode, text, 320, 320)
        height, width = bitmap.shape
        image = Image.frombytes("L", (width, height), bytes(bitmap))
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()

    def sample_pdf_path(self, filename: str) -> Path:
        path = Path("C:/Users/Administrator/Downloads") / filename
        if not path.exists():
            self.skipTest(f"样本 PDF 不存在：{path}")
        return path

    def test_default_admin_login(self) -> None:
        response = self.client.post("/api/auth/login", json={"username": "管理员1", "password": "123456"})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["ok"])

    def test_add_duplicate_invoice_rejected(self) -> None:
        self.login()
        invoice = {
            "invoice_code": "2026-06-24-001",
            "original_invoice_code": "011002300111",
            "invoice_number": "12345678",
            "issue_date": "2024-05-21",
            "buyer_name": "购买方",
            "seller_name": "销售方",
            "amount": "100.00",
            "tax_amount": "13.00",
            "total_amount": "113.00",
            "remark": "",
        }
        first = self.client.post("/api/invoices", json=invoice)
        self.assertEqual(first.status_code, 200)
        duplicate_invoice = {
            **invoice,
            "invoice_code": "2026-06-24-002",
            "original_invoice_code": "022002300222",
            "issue_date": "2025-06-01",
            "total_amount": "999.00",
            "allow_duplicate": True,
        }
        duplicate = self.client.post("/api/invoices", json=duplicate_invoice)
        self.assertEqual(duplicate.status_code, 409)
        self.assertFalse(duplicate.get_json()["ok"])
        self.assertTrue(duplicate.get_json()["data"]["duplicates"])

    def test_invoice_number_duplicate_rejected_across_companies(self) -> None:
        self.login()
        saved = self.client.post("/api/invoices", json=self.sample_invoice())
        self.assertEqual(saved.status_code, 200)

        created = self.client.post("/api/companies", json={"name": "公司B"})
        self.assertEqual(created.status_code, 200)

        duplicate_invoice = self.sample_invoice(
            code="2026-06-24-009",
            number=self.sample_invoice()["invoice_number"],
        )
        duplicate_invoice["original_invoice_code"] = "022002300222"
        duplicate_invoice["issue_date"] = "2026-06-01"
        duplicate_invoice["total_amount"] = "999.00"
        duplicate = self.client.post("/api/invoices", json=duplicate_invoice)
        self.assertEqual(duplicate.status_code, 409)
        data = duplicate.get_json()["data"]
        self.assertEqual(data["duplicates"][0]["company_name"], DEFAULT_COMPANY_NAME)

    def test_company_scope_and_join_same_company_name(self) -> None:
        self.login()
        created = self.client.post("/api/companies", json={"name": "公司A"})
        self.assertEqual(created.status_code, 200)
        company = created.get_json()["data"]["company"]

        saved = self.client.post("/api/invoices", json=self.sample_invoice())
        self.assertEqual(saved.status_code, 200)

        self.client.post("/api/auth/logout")
        self.store.add_admin(DEFAULT_ADMIN, DEFAULT_PASSWORD, "财务B", "123456")
        self.login("财务B", "123456")

        no_company = self.client.get("/api/invoices")
        self.assertEqual(no_company.status_code, 409)

        joined = self.client.post("/api/companies", json={"name": "公司A"})
        self.assertEqual(joined.status_code, 200)
        self.assertEqual(joined.get_json()["data"]["company"]["id"], company["id"])

        records = self.client.get("/api/invoices")
        self.assertEqual(records.status_code, 200)
        self.assertEqual(records.get_json()["data"]["total"], 1)

    def test_duplicate_check_scans_all_companies(self) -> None:
        self.login()
        saved = self.client.post("/api/invoices", json=self.sample_invoice())
        self.assertEqual(saved.status_code, 200)

        created = self.client.post("/api/companies", json={"name": "公司B"})
        self.assertEqual(created.status_code, 200)

        duplicate = self.client.post("/api/invoices/check", json=self.sample_invoice(code="2026-06-24-009"))
        self.assertEqual(duplicate.status_code, 200)
        data = duplicate.get_json()["data"]
        self.assertTrue(data["exists"])
        self.assertEqual(data["duplicates"][0]["company_name"], DEFAULT_COMPANY_NAME)

    def test_questionable_records_are_company_scoped_and_clearable(self) -> None:
        self.login()
        created = self.client.post("/api/questionable", json={"file_name": "a.pdf", "reason": "重复发票", "note": "文件夹内重复"})
        self.assertEqual(created.status_code, 200)

        listed = self.client.get("/api/questionable")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(len(listed.get_json()["data"]["items"]), 1)

        cleared = self.client.post("/api/questionable/clear", json={"scope": "today"})
        self.assertEqual(cleared.status_code, 200)
        self.assertEqual(cleared.get_json()["data"]["deleted"], 1)

    def test_invoice_pagination_ids_and_batch_delete(self) -> None:
        self.login()
        for index in range(25):
            self.store.add_invoice(
                self.sample_invoice(code=f"2026-06-24-{index + 1:03d}", number=f"900000{index:02d}"),
                actor=DEFAULT_ADMIN,
            )

        page = self.client.get("/api/invoices?page=2&page_size=20")
        self.assertEqual(page.status_code, 200)
        page_data = page.get_json()["data"]
        self.assertEqual(page_data["total"], 25)
        self.assertEqual(page_data["page"], 2)
        self.assertEqual(len(page_data["items"]), 5)

        ids_response = self.client.get("/api/invoices/ids")
        ids = ids_response.get_json()["data"]["ids"]
        self.assertEqual(len(ids), 25)

        deleted = self.client.post("/api/invoices/batch-delete", json={"ids": ids[:3]})
        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(deleted.get_json()["data"]["deleted"], 3)
        self.assertEqual(self.client.get("/api/invoices").get_json()["data"]["total"], 22)

    def test_last_admin_can_remove_company_without_deleting_database_and_rejoin(self) -> None:
        self.login()
        created = self.client.post("/api/companies", json={"name": "可恢复公司", "remark": "保留数据"})
        self.assertEqual(created.status_code, 200)
        company = created.get_json()["data"]["company"]
        self.assertEqual(company["remark"], "保留数据")

        saved = self.client.post("/api/invoices", json=self.sample_invoice())
        self.assertEqual(saved.status_code, 200)

        removed = self.client.post("/api/companies/delete", json={"company_id": company["id"], "delete_database": False})
        self.assertEqual(removed.status_code, 200)
        self.assertFalse(removed.get_json()["data"]["database_deleted"])

        rejoined = self.client.post("/api/companies", json={"name": "可恢复公司"})
        self.assertEqual(rejoined.status_code, 200)
        self.assertEqual(rejoined.get_json()["data"]["company"]["id"], company["id"])

        records = self.client.get("/api/invoices")
        self.assertEqual(records.status_code, 200)
        self.assertEqual(records.get_json()["data"]["total"], 1)

    def test_expired_records_are_purged_by_created_at(self) -> None:
        self.login()
        self.store.add_invoice(self.sample_invoice(), actor=DEFAULT_ADMIN)
        path = self.store.company_invoices_path(DEFAULT_COMPANY_ID)
        payload = self.store._read_json(path, {"items": []})
        payload["items"][0]["created_at"] = (datetime.now() - timedelta(days=600)).strftime("%Y-%m-%d %H:%M:%S")
        self.store._write_json(path, payload)

        response = self.client.get("/api/invoices")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"]["total"], 0)

    def test_export_filters_by_created_at_range(self) -> None:
        self.login()
        self.store.add_invoice(self.sample_invoice(code="2026-06-24-001", number="10000001"), actor=DEFAULT_ADMIN)
        self.store.add_invoice(self.sample_invoice(code="2026-06-24-002", number="10000002"), actor=DEFAULT_ADMIN)
        path = self.store.company_invoices_path(DEFAULT_COMPANY_ID)
        payload = self.store._read_json(path, {"items": []})
        payload["items"][0]["created_at"] = "2026-01-10 10:00:00"
        payload["items"][1]["created_at"] = "2026-02-10 10:00:00"
        self.store._write_json(path, payload)

        response = self.client.get("/api/excel/export?start=2026-02-01&end=2026-02-28")
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        try:
            rows = list(wb.active.iter_rows(values_only=True))
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][2], "10000002")
        finally:
            wb.close()
            response.close()

    def test_parse_invoice_text_recognizes_issue_date(self) -> None:
        parsed = parse_invoice_text(
            "发票代码：011002300111\n发票号码：12345678\n开票日期：2026年6月24日\n价税合计：113.00"
        )
        self.assertEqual(parsed["issue_date"], "2026-06-24")

    def test_parse_sample_scanned_pdf_reads_number_date_and_total(self) -> None:
        parsed = parse_invoice_file(self.sample_pdf_path("390167bc97621a6860458e46871ad47b.pdf"))
        self.assertEqual(parsed["invoice_number"], "26132000000650798086")
        self.assertEqual(parsed["issue_date"], "2026-03-09")
        self.assertEqual(parsed["total_amount"], "329.00")
        self.assertIn("二维码内容", parsed.get("raw_text", ""))

    def test_parse_sample_railway_pdf_reads_number_date_and_total(self) -> None:
        parsed = parse_invoice_file(
            self.sample_pdf_path("dzfp_26419165773006463048_京广铁路客运专线河南有限责任公司_202606.pdf")
        )
        self.assertEqual(parsed["invoice_number"], "26419165773006463048")
        self.assertEqual(parsed["issue_date"], "2026-06-25")
        self.assertEqual(parsed["total_amount"], "921.00")
        self.assertIn("二维码内容", parsed.get("raw_text", ""))

    def test_next_daily_invoice_code_counts_existing_records(self) -> None:
        self.login()
        self.store.add_invoice(
            {
                "invoice_code": "2026-06-24-001",
                "original_invoice_code": "011002300111",
                "invoice_number": "12345678",
                "issue_date": "2024-05-21",
                "buyer_name": "购买方",
                "seller_name": "销售方",
                "amount": "100.00",
                "tax_amount": "13.00",
                "total_amount": "113.00",
                "remark": "",
            },
            actor="管理员1",
        )
        self.assertEqual(self.store.next_daily_invoice_code("2026-06-24"), "2026-06-24-002")

    def test_fuzzy_digit_search(self) -> None:
        self.login()
        self.store.add_invoice(
            {
                "invoice_code": "2026-06-24-001",
                "original_invoice_code": "011002300111",
                "invoice_number": "987654321",
                "issue_date": "2024-05-21",
                "buyer_name": "购买方",
                "seller_name": "销售方",
                "amount": "100.00",
                "tax_amount": "13.00",
                "total_amount": "113.00",
                "remark": "",
            },
            actor="管理员1",
        )
        response = self.client.get("/api/invoices?q=7654")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["data"]["total"], 1)

    def test_excel_template_can_be_read(self) -> None:
        path = create_template(Path(self.tmp.name) / "template.xlsx")
        wb = load_workbook(path)
        try:
            self.assertIn("发票登记模板", wb.sheetnames)
        finally:
            wb.close()
        rows = read_invoice_rows(path, self.store)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["invoice"]["invoice_number"], "12345678")

    def test_upload_accepts_pdf_file(self) -> None:
        self.login()
        response = self.client.post(
            "/api/files/parse",
            data={"file": (io.BytesIO(self.make_pdf_bytes()), "invoice.pdf")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["data"]["source_file"], "invoice.pdf")
        self.assertRegex(payload["data"]["invoice_code"], r"^\d{4}-\d{2}-\d{2}-\d{3}$")

    def test_upload_accepts_jpg_file(self) -> None:
        self.login()
        response = self.client.post(
            "/api/files/parse",
            data={"file": (io.BytesIO(self.make_jpg_bytes()), "invoice.jpg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["data"]["source_file"], "invoice.jpg")

    def test_legacy_upload_route_accepts_jpg_file(self) -> None:
        self.login()
        response = self.client.post(
            "/api/pdf/parse",
            data={"file": (io.BytesIO(self.make_jpg_bytes()), "invoice.jpg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["data"]["source_file"], "invoice.jpg")

    def test_qrcode_parse_fills_invoice_fields(self) -> None:
        self.login()
        response = self.client.post(
            "/api/qrcode/parse",
            json={"text": "01,04,011002300111,12345678,113.00,20240521,12345678901234567890,ABCD,"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        data = payload["data"]
        self.assertTrue(re.match(r"^\d{4}-\d{2}-\d{2}-\d{3}$", data["invoice_code"]))
        self.assertEqual(data["original_invoice_code"], "011002300111")
        self.assertEqual(data["invoice_number"], "12345678")
        self.assertEqual(data["issue_date"], "2024-05-21")
        self.assertEqual(data["total_amount"], "113.00")

    def test_qrcode_image_upload_fills_invoice_fields(self) -> None:
        self.login()
        qr_text = "01,04,011002300111,12345678,113.00,20240521,12345678901234567890,ABCD,"
        response = self.client.post(
            "/api/qrcode/image",
            data={"file": (io.BytesIO(self.make_qr_png_bytes(qr_text)), "invoice-qr.png")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        data = payload["data"]
        self.assertTrue(re.match(r"^\d{4}-\d{2}-\d{2}-\d{3}$", data["invoice_code"]))
        self.assertEqual(data["original_invoice_code"], "011002300111")
        self.assertEqual(data["invoice_number"], "12345678")
        self.assertEqual(data["issue_date"], "2024-05-21")
        self.assertEqual(data["total_amount"], "113.00")

    def test_upload_route_accepts_trailing_slash(self) -> None:
        self.login()
        response = self.client.post(
            "/api/files/parse/",
            data={"file": (io.BytesIO(self.make_pdf_bytes()), "invoice.pdf")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["data"]["source_file"], "invoice.pdf")

    def test_upload_rejects_unsupported_file(self) -> None:
        self.login()
        response = self.client.post(
            "/api/files/parse",
            data={"file": (io.BytesIO(b"not an invoice"), "invoice.txt")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.get_json()["ok"])

    def test_method_not_allowed_returns_json_405(self) -> None:
        response = self.client.post("/api/excel/template")
        self.assertEqual(response.status_code, 405)
        payload = response.get_json()
        self.assertFalse(payload["ok"])
        self.assertIn("not allowed", payload["message"].lower())

    def test_environment_check_passes_in_development_workspace(self) -> None:
        issues = check_runtime_environment()
        errors = [issue for issue in issues if issue.get("level") == "error"]
        self.assertEqual(errors, [], format_environment_report(issues))


if __name__ == "__main__":
    unittest.main()
