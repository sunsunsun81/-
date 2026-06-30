from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from .config import exports_dir
from .data_store import JsonStore, normalize_key_part

HEADERS = [
    "发票编码",
    "原始发票代码",
    "发票号码",
    "开票日期",
    "购买方名称",
    "销售方名称",
    "金额",
    "税额",
    "价税合计",
    "备注",
]

FIELD_MAP = {
    "发票编码": "invoice_code",
    "原始发票代码": "original_invoice_code",
    "发票号码": "invoice_number",
    "开票日期": "issue_date",
    "购买方名称": "buyer_name",
    "销售方名称": "seller_name",
    "金额": "amount",
    "税额": "tax_amount",
    "价税合计": "total_amount",
    "备注": "remark",
}


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D7DEE8")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    widths = [18, 16, 16, 14, 28, 28, 14, 14, 14, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def create_template(path: Optional[Path] = None) -> Path:
    output = path or exports_dir() / "发票导入模板.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "发票登记模板"
    ws.append(HEADERS)
    ws.append(["2026-06-24-001", "011002300111", "12345678", "2024-05-21", "示例购买方有限公司", "示例销售方有限公司", 3540.71, 459.29, 4000.00, "示例行，可删除"])
    _style_sheet(ws)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    wb.save(output)
    return output


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def read_invoice_rows(path: Path, store: JsonStore, company_id: Optional[str] = None) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header_row = [_cell_to_text(value) for value in rows[0]]
        header_index = {name: idx for idx, name in enumerate(header_row)}
        missing_headers = [name for name in HEADERS[2:9] if name not in header_index]
        if missing_headers:
            raise ValueError("Excel 缺少表头：" + "、".join(missing_headers))

        result: List[Dict[str, Any]] = []
        next_generated_code = store.next_daily_invoice_code(company_id)
        generated_match = next_generated_code.rsplit("-", 1)
        generated_prefix = generated_match[0] if len(generated_match) == 2 else datetime.now().strftime("%Y-%m-%d")
        generated_number = int(generated_match[1]) if len(generated_match) == 2 and generated_match[1].isdigit() else 1
        for excel_row_number, row in enumerate(rows[1:], start=2):
            if not any(_cell_to_text(value) for value in row):
                continue
            invoice: Dict[str, Any] = {}
            for header, field in FIELD_MAP.items():
                idx = header_index.get(header)
                invoice[field] = _cell_to_text(row[idx]) if idx is not None and idx < len(row) else ""
            if not invoice.get("original_invoice_code") and "发票代码" in header_index:
                idx = header_index["发票代码"]
                invoice["original_invoice_code"] = _cell_to_text(row[idx]) if idx < len(row) else ""
            if not invoice.get("invoice_code"):
                invoice["invoice_code"] = f"{generated_prefix}-{generated_number:03d}"
                generated_number += 1
            invoice["invoice_code"] = str(invoice.get("invoice_code", "")).strip()
            invoice["original_invoice_code"] = normalize_key_part(invoice.get("original_invoice_code"))
            invoice["invoice_number"] = normalize_key_part(invoice.get("invoice_number"))
            duplicates = store.find_duplicates_all_companies(invoice)
            duplicate = duplicates[0]["invoice"] if duplicates else None
            warnings = []
            if not invoice["invoice_number"]:
                warnings.append("缺少发票号码")
            if duplicate:
                warnings.append("历史数据中已存在相同发票号码")
            result.append(
                {
                    "row_number": excel_row_number,
                    "invoice": invoice,
                    "duplicate": duplicate,
                    "duplicates": duplicates,
                    "warnings": warnings,
                    "suggested_action": "question" if warnings else "save",
                }
            )
        return result
    finally:
        wb.close()


def export_invoices(records: Iterable[Dict[str, Any]], filename: str = "发票历史数据.xlsx") -> Path:
    output = exports_dir() / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "历史数据"
    ws.append(HEADERS + ["登记人", "登记时间"])
    for item in records:
        ws.append(
            [
                item.get("invoice_code", ""),
                item.get("original_invoice_code", ""),
                item.get("invoice_number", ""),
                item.get("issue_date", ""),
                item.get("buyer_name", ""),
                item.get("seller_name", ""),
                item.get("amount", ""),
                item.get("tax_amount", ""),
                item.get("total_amount", ""),
                item.get("remark", ""),
                item.get("created_by", ""),
                item.get("created_at", ""),
            ]
        )
    _style_sheet(ws)
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 20
    wb.save(output)
    return output


def export_questionable(rows: Iterable[Dict[str, Any]]) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = exports_dir() / f"疑惑发票_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "疑惑记录"
    ws.append(["Excel行号"] + HEADERS + ["疑惑原因", "处理动作"])
    for row in rows:
        invoice = row.get("invoice", {})
        ws.append(
            [
                row.get("row_number", ""),
                invoice.get("invoice_code", ""),
                invoice.get("original_invoice_code", ""),
                invoice.get("invoice_number", ""),
                invoice.get("issue_date", ""),
                invoice.get("buyer_name", ""),
                invoice.get("seller_name", ""),
                invoice.get("amount", ""),
                invoice.get("tax_amount", ""),
                invoice.get("total_amount", ""),
                invoice.get("remark", ""),
                "；".join(row.get("warnings", [])),
                row.get("action", ""),
            ]
        )
    _style_sheet(ws)
    ws.column_dimensions["K"].width = 32
    ws.column_dimensions["L"].width = 42
    ws.column_dimensions["M"].width = 14
    wb.save(output)
    return output
